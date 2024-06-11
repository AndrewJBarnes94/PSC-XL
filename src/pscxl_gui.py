import sqlite3
from collections import defaultdict
import tkinter as tk
import openpyxl
from tkinter import ttk, filedialog, messagebox
import threading
import logging
from pscxl_database import PSCXL_Database

# Configure logging
logging.basicConfig(level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s', filename='pscxl_log.txt', filemode='w')

class PSCXL_GUI:
    def __init__(self, root):
        self.root = root
        self.root.title("PSCXL")
        logging.debug('Initialized PSCXL_GUI')

        # Add a button to open a new Excel file
        self.new_file_button = tk.Button(root, text="New Excel File", command=self.open_file)
        self.new_file_button.pack(pady=10)

        # Add a button to create a new Excel document from the database
        self.export_button = tk.Button(root, text="Export to Excel", command=self.export_to_excel)
        self.export_button.pack(pady=10)

        # Create a dropdown menu for workbook names (initially with a default option)
        self.workbook_selector = ttk.Combobox(root, values=["--Select--"])
        self.workbook_selector.bind("<<ComboboxSelected>>", self.update_sheet_selector)
        self.workbook_selector.current(0)
        self.workbook_selector.pack(pady=10)

        # Create a dropdown menu for sheet names (initially with a default option)
        self.sheet_selector = ttk.Combobox(root, values=["--Select--"])
        self.sheet_selector.bind("<<ComboboxSelected>>", self.update_table)
        self.sheet_selector.current(0)
        self.sheet_selector.pack(pady=10)

        # Add a Treeview widget to display the database contents
        self.db_tree = ttk.Treeview(root)

        # Define columns for the database Treeview
        self.db_tree['columns'] = ('Workbook', 'Sheet', 'Value', 'Quantity', 'Stacked')

        # Format columns
        self.db_tree.column("#0", width=0, stretch=tk.NO)  # Hide the first column
        self.db_tree.column('Workbook', anchor=tk.W, width=150)
        self.db_tree.column('Sheet', anchor=tk.W, width=150)
        self.db_tree.column('Value', anchor=tk.W, width=250)
        self.db_tree.column('Quantity', anchor=tk.CENTER, width=80)
        self.db_tree.column('Stacked', anchor=tk.CENTER, width=80)

        # Create headings
        self.db_tree.heading("#0", text="", anchor=tk.W)
        self.db_tree.heading('Workbook', text='Workbook', anchor=tk.W)
        self.db_tree.heading('Sheet', text='Sheet', anchor=tk.W)
        self.db_tree.heading('Value', text='Value', anchor=tk.W)
        self.db_tree.heading('Quantity', text='Quantity', anchor=tk.CENTER)
        self.db_tree.heading('Stacked', text='Stacked', anchor=tk.CENTER)

        # Pack Treeview widget
        self.db_tree.pack(pady=20)

        # Bind right-click menu to Treeview
        self.db_tree.bind("<Button-3>", self.show_context_menu)

        # Create a context menu for right-click actions
        self.context_menu = tk.Menu(root, tearoff=0)
        self.context_menu.add_command(label="Edit Row", command=self.edit_row)
        self.context_menu.add_command(label="Add Row", command=self.add_row)

        # Initialize SQLite database
        self.db = PSCXL_Database()

        # Initialize SQLite database cursor
        self.cursor = self.db.cursor

        # Read the database to populate the workbook selector
        self.read_database()

    def disable_buttons(self):
        self.new_file_button.config(state=tk.DISABLED)
        self.export_button.config(state=tk.DISABLED)

    def enable_buttons(self):
        self.new_file_button.config(state=tk.NORMAL)
        self.export_button.config(state=tk.NORMAL)

    def open_file(self):
        logging.debug('open_file called')
        # Open a file dialog to select the Excel file
        file_path = filedialog.askopenfilename(
            title="Select Excel File",
            filetypes=(("Excel files", "*.xlsx"), ("All files", "*.*"))
        )
        if file_path:
            logging.debug(f'Selected file: {file_path}')
            self.workbook_name = file_path
            self.wb = openpyxl.load_workbook(self.workbook_name)
            self.sheet_names = self.wb.sheetnames

            # Store data for all sheets in the database
            workbook_base_name = self.workbook_name.split('/')[-1].split('.')[0]

            # Clear existing data for this workbook in the database
            self.db.clear_workbook_data(workbook_base_name)

            for sheet_name in self.sheet_names:
                sheet = self.wb[sheet_name]
                sheet_data = self.read_sheet(sheet)
                duplicate_counts = self.count_duplicates(sheet_data)
                for value, count in duplicate_counts.items():
                    # Normalize the value
                    normalized_value = str(value).strip()
                    # Check if value is a string before checking for spaces
                    if isinstance(value, str) and " " in normalized_value:
                        stacked = 1  # Assume values with spaces are stacked
                    else:
                        stacked = 0
                    self.db.insert_data(workbook_base_name, sheet_name, normalized_value, count, stacked)
            self.read_database()

    def read_sheet(self, sheet):
        logging.debug(f'Reading sheet: {sheet.title}')
        data = []
        for row in sheet.iter_rows(values_only=True):  # Read all rows
            data.append(list(row))  # Convert each row to a list and append to data
        return data

    def count_duplicates(self, data):
        logging.debug('Counting duplicates')
        counter = defaultdict(int)
        for row in data:
            for value in row:
                if value is not None:  # Skip empty cells
                    counter[value] += 1
        return counter

    def read_database(self):
        logging.debug('Reading database')
        # Get unique workbook names and update the workbook selector
        workbooks = ["--Select--"] + self.db.fetch_workbooks()
        self.workbook_selector['values'] = workbooks
        self.workbook_selector.current(0)
        self.update_sheet_selector()

    def update_sheet_selector(self, event=None):
        logging.debug('Updating sheet selector')
        # Get the selected workbook name
        selected_workbook = self.workbook_selector.get()

        if selected_workbook == "--Select--":
            self.sheet_selector['values'] = ["--Select--"]
            self.sheet_selector.current(0)
            self.clear_table()
            return

        # Get unique sheet names for the selected workbook and update the sheet selector
        sheets = ["--Select--"] + self.db.fetch_sheets(selected_workbook)
        self.sheet_selector['values'] = sheets
        self.sheet_selector.current(0)
        self.clear_table()

    def update_table(self, event=None):
        logging.debug('Updating table')
        # Clear the current table
        self.clear_table()

        # Get the selected workbook and sheet name
        selected_workbook = self.workbook_selector.get()
        selected_sheet = self.sheet_selector.get()

        if selected_workbook == "--Select--" or selected_sheet == "--Select--":
            return

        # Query the database for the selected workbook and sheet
        rows = self.db.fetch_data(selected_workbook, selected_sheet)

        # Insert data into the database Treeview widget
        for row in rows:
            display_row = list(row)
            if isinstance(display_row[2], str):
                display_row[2] = display_row[2].replace(" " * 10, " ")  # Show single value without extra spaces
            self.db_tree.insert(parent='', index='end', values=display_row)

    def clear_table(self):
        logging.debug('Clearing table')
        for item in self.db_tree.get_children():
            self.db_tree.delete(item)

    def show_context_menu(self, event):
        logging.debug('Showing context menu')
        # Show context menu on right-click
        if not self.db_tree.selection():
            messagebox.showinfo("Info", "Please select a row first.")
        else:
            try:
                self.context_menu.tk_popup(event.x_root, event.y_root)
            finally:
                self.context_menu.grab_release()

    def edit_row(self):
        logging.debug('Editing row')
        # Get the selected row
        selected_item = self.db_tree.selection()
        if not selected_item:
            messagebox.showinfo("Info", "Please select a row first.")
            return
        item = self.db_tree.item(selected_item)

        # Get current values
        current_values = item['values']
        new_values = self.edit_popup(current_values)
        if new_values:
            # Update the database
            self.db.update_data(current_values[0], current_values[1], current_values[2], new_values[2], new_values[3], new_values[4])
            self.update_table()

    def add_row(self):
        logging.debug('Adding row')
        # Get the selected workbook and sheet name
        selected_workbook = self.workbook_selector.get()
        selected_sheet = self.sheet_selector.get()

        if selected_workbook == "--Select--" or selected_sheet == "--Select--":
            messagebox.showinfo("Info", "Please select a workbook and sheet first.")
            return

        # Get new values
        new_values = self.edit_popup([selected_workbook, selected_sheet, "", 0, 0], new_row=True)
        if new_values:
            # Insert into the database
            self.db.insert_data(new_values[0], new_values[1], new_values[2], new_values[3], new_values[4])
            self.update_table()

    def edit_popup(self, values, new_row=False):
        logging.debug('Opening edit popup')
        popup = tk.Toplevel()
        popup.title("Edit Row" if not new_row else "Add Row")

        tk.Label(popup, text="Workbook:").grid(row=0, column=0, padx=10, pady=10)
        tk.Label(popup, text=values[0]).grid(row=0, column=1, padx=10, pady=10)

        tk.Label(popup, text="Sheet:").grid(row=1, column=0, padx=10, pady=10)
        tk.Label(popup, text=values[1]).grid(row=1, column=1, padx=10, pady=10)

        tk.Label(popup, text="Value:").grid(row=2, column=0, padx=10, pady=10)
        value_entry = tk.Entry(popup)
        value_entry.grid(row=2, column=1, padx=10, pady=10)
        value_entry.insert(0, values[2])

        tk.Label(popup, text="Quantity:").grid(row=3, column=0, padx=10, pady=10)
        quantity_entry = tk.Entry(popup)
        quantity_entry.grid(row=3, column=1, padx=10, pady=10)
        quantity_entry.insert(0, values[3])

        stacked_var = tk.IntVar(value=values[4])
        stacked_check = tk.Checkbutton(popup, text="Stacked", variable=stacked_var)
        stacked_check.grid(row=4, column=0, columnspan=2, pady=10)

        def on_submit():
            logging.debug('Submitting edit popup')
            popup.new_values = [values[0], values[1], value_entry.get(), int(quantity_entry.get()), stacked_var.get()]
            popup.destroy()

        submit_button = tk.Button(popup, text="Submit", command=on_submit)
        submit_button.grid(row=5, column=0, columnspan=2, pady=10)

        popup.transient(self.root)
        popup.grab_set()
        self.root.wait_window(popup)

        return getattr(popup, 'new_values', None)

    def export_to_excel(self):
        logging.debug('Exporting to Excel')
        # Open a file dialog to select the save location for the new Excel file
        save_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=(("Excel files", "*.xlsx"), ("All files", "*.*")),
            title="Save Excel File"
        )
        if save_path:
            self.disable_buttons()
            threading.Thread(target=self.run_export_to_excel, args=(save_path,)).start()

    def run_export_to_excel(self, save_path):
        try:
            new_wb = openpyxl.Workbook()
            new_wb.remove(new_wb.active)  # Remove the default sheet

            workbooks = self.db.fetch_workbooks()
            for workbook_name in workbooks:
                sheets = self.db.fetch_sheets(workbook_name)
                for sheet_name in sheets:
                    new_ws = new_wb.create_sheet(title=sheet_name)
                    new_ws.append(['Value', 'Quantity'])

                    rows = self.db.fetch_data(workbook_name, sheet_name)
                    for row in rows:
                        value, quantity, stacked = row
                        for _ in range(quantity):
                            if stacked:
                                value = value.replace(" ", " " * 10)  # Ensure 10 spaces between stacked items
                            new_ws.append([value])

            new_wb.save(save_path)
            logging.debug(f'Data successfully exported to {save_path}')
            messagebox.showinfo("Info", "Data successfully exported to Excel")
        except Exception as e:
            logging.error(f'Error saving Excel file: {e}')
            messagebox.showerror("Error", f"Failed to export data to Excel: {e}")
        finally:
            self.enable_buttons()

if __name__ == "__main__":
    logging.debug('Starting application')
    root = tk.Tk()
    app = PSCXL_GUI(root)
    root.mainloop()
