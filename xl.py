import openpyxl
import json
from collections import defaultdict

# Define the workbook name
workbook_name = "24-0005E2.xlsx"

# Open the workbook
wb = openpyxl.load_workbook(workbook_name)

# Function to read all cell values from a sheet and store them in a structured format
def read_sheet(sheet):
    data = []
    for row in sheet.iter_rows(values_only=True):  # Read all rows
        data.append(list(row))  # Convert each row to a list and append to data
    return data

# Function to count duplicate values in a list of lists
def count_duplicates(data):
    counter = defaultdict(int)
    for row in data:
        for value in row:
            if value is not None:  # Skip empty cells
                counter[value] += 1
    duplicates = {key: count for key, count in counter.items() if count > 1}
    return duplicates

# Dictionary to hold data and duplicate counts from each sheet
all_data = {}
duplicate_counts = {}

# Get all sheet names dynamically
sheet_names = wb.sheetnames

# Loop through each sheet and read data
for sheet_name in sheet_names:
    sheet = wb[sheet_name]
    sheet_data = read_sheet(sheet)
    all_data[sheet_name] = sheet_data
    duplicate_counts[sheet_name] = count_duplicates(sheet_data)

# Print the structured data
'''for sheet_name, data in all_data.items():
    print(f"\nData from sheet {sheet_name}:")
    for row in data:
        print(row)'''

# Print duplicate counts
for sheet_name, counts in duplicate_counts.items():
    print(f"\nDuplicate counts for sheet {sheet_name}:")
    for value, count in counts.items():
        print(f"Value: {value}, Count: {count}")

# Optionally, save the structured data and duplicate counts to JSON files
with open('workbook_data.json', 'w') as json_file:
    json.dump(all_data, json_file, indent=4)

with open('duplicate_counts.json', 'w') as json_file:
    json.dump(duplicate_counts, json_file, indent=4)
