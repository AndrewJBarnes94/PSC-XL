import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import openpyxl
from collections import defaultdict
import sqlite3

class PSCXL:
    def __init__(self, root):
        self.root = root
        self.root.title("PSCXL")

        # Add a button to open a new Excel file
        self.new_file_button = tk.Button(root, text="New Excel File", command=self.open_file)
        self.new_file_button.pack(pady=10)

        # Add a button to create a new Excel document from the database
        self.export_button = tk.Button(root, text="Export to Excel", command=self.export_to_excel)
        self.export_button.pack(pady=10)

        # Create a dropdown menu for workbook names (initially with a default option)
        self.workbook_selector = ttk.Combobox(root, values=["--Select--"])
        self.workbook_selector.bind("<<ComboboxSelected>>", self.update_sheet_selector)
        self.workbook_selector.current(0)
        self.workbook_selector.pack(pady=10)

        # Create a dropdown menu for sheet names (initially with a default option)
        self.sheet_selector = ttk.Combobox(root, values=["--Select--"])
        self.sheet_selector.bind("<<ComboboxSelected>>", self.update_table)
        self.sheet_selector.current(0)
        self.sheet_selector.pack(pady=10)

        # Add a Treeview widget to display the database contents
        self.db_tree = ttk.Treeview(root)

        # Define columns for the database Treeview
        self.db_tree['columns'] = ('Workbook', 'Sheet', 'Value', 'Quantity', 'Stacked')

        # Format columns
        self.db_tree.column("#0", width=0, stretch=tk.NO)  # Hide the first column
        self.db_tree.column('Workbook', anchor=tk.W, width=150)
        self.db_tree.column('Sheet', anchor=tk.W, width=150)
        self.db_tree.column('Value', anchor=tk.W, width=250)
        self.db_tree.column('Quantity', anchor=tk.CENTER, width=80)
        self.db_tree.column('Stacked', anchor=tk.CENTER, width=80)

        # Create headings
        self.db_tree.heading("#0", text="", anchor=tk.W)
        self.db_tree.heading('Workbook', text='Workbook', anchor=tk.W)
        self.db_tree.heading('Sheet', text='Sheet', anchor=tk.W)
        self.db_tree.heading('Value', text='Value', anchor=tk.W)
        self.db_tree.heading('Quantity', text='Quantity', anchor=tk.CENTER)
        self.db_tree.heading('Stacked', text='Stacked', anchor=tk.CENTER)

        # Pack Treeview widget
        self.db_tree.pack(pady=20)

        # Bind right-click menu to Treeview
        self.db_tree.bind("<Button-3>", self.show_context_menu)

        # Create a context menu for right-click actions
        self.context_menu = tk.Menu(root, tearoff=0)
        self.context_menu.add_command(label="Edit Row", command=self.edit_row)
        self.context_menu.add_command(label="Add Row", command=self.add_row)

        # Initialize SQLite database
        self.conn = sqlite3.connect('pscxl.db')
        self.cursor = self.conn.cursor()
        self.cursor.execute('''
            CREATE TABLE IF NOT EXISTS workbooks (
                workbook_name TEXT, 
                sheet_name TEXT, 
                value TEXT, 
                quantity INTEGER,
                stacked INTEGER,
                PRIMARY KEY (workbook_name, sheet_name, value)
            )
        ''')
        self.conn.commit()

        # Read the database to populate the workbook selector
        self.read_database()

    def open_file(self):
        # Open a file dialog to select the Excel file
        file_path = filedialog.askopenfilename(
            title="Select Excel File",
            filetypes=(("Excel files", "*.xlsx"), ("All files", "*.*"))
        )
        if file_path:
            self.workbook_name = file_path
            self.wb = openpyxl.load_workbook(self.workbook_name)
            self.sheet_names = self.wb.sheetnames

            # Store data for all sheets in the database
            workbook_base_name = self.workbook_name.split('/')[-1].split('.')[0]
            for sheet_name in self.sheet_names:
                sheet = self.wb[sheet_name]
                sheet_data = self.read_sheet(sheet)
                duplicate_counts = self.count_duplicates(sheet_data)
                for value, count in duplicate_counts.items():
                    stacked = 1 if " " in value else 0  # Assume values with spaces are stacked
                    self.cursor.execute('''
                        INSERT OR IGNORE INTO workbooks (workbook_name, sheet_name, value, quantity, stacked) 
                        VALUES (?, ?, ?, ?, ?)
                    ''', (workbook_base_name, sheet_name, value.strip(), count, stacked))
            self.conn.commit()
            self.read_database()

    def read_sheet(self, sheet):
        data = []
        for row in sheet.iter_rows(values_only=True):  # Read all rows
            data.append(list(row))  # Convert each row to a list and append to data
        return data

    def count_duplicates(self, data):
        counter = defaultdict(int)
        for row in data:
            for value in row:
                if value is not None:  # Skip empty cells
                    counter[value] += 1
        return counter

    def read_database(self):
        # Get unique workbook names and update the workbook selector
        self.cursor.execute('SELECT DISTINCT workbook_name FROM workbooks')
        workbooks = ["--Select--"] + [row[0] for row in self.cursor.fetchall()]
        self.workbook_selector['values'] = workbooks
        self.workbook_selector.current(0)
        self.update_sheet_selector()

    def update_sheet_selector(self, event=None):
        # Get the selected workbook name
        selected_workbook = self.workbook_selector.get()

        if selected_workbook == "--Select--":
            self.sheet_selector['values'] = ["--Select--"]
            self.sheet_selector.current(0)
            self.clear_table()
            return

        # Get unique sheet names for the selected workbook and update the sheet selector
        self.cursor.execute('SELECT DISTINCT sheet_name FROM workbooks WHERE workbook_name = ?', (selected_workbook,))
        sheets = ["--Select--"] + [row[0] for row in self.cursor.fetchall()]
        self.sheet_selector['values'] = sheets
        self.sheet_selector.current(0)
        self.clear_table()

    def update_table(self, event=None):
        # Clear the current table
        self.clear_table()

        # Get the selected workbook and sheet name
        selected_workbook = self.workbook_selector.get()
        selected_sheet = self.sheet_selector.get()

        if selected_workbook == "--Select--" or selected_sheet == "--Select--":
            return

        # Query the database for the selected workbook and sheet
        self.cursor.execute('SELECT * FROM workbooks WHERE workbook_name = ? AND sheet_name = ?', 
                            (selected_workbook, selected_sheet))
        rows = self.cursor.fetchall()

        # Insert data into the database Treeview widget
        for row in rows:
            display_row = list(row)
            display_row[2] = display_row[2].replace(" " * 10, " ")  # Show single value without extra spaces
            self.db_tree.insert(parent='', index='end', values=display_row)

    def clear_table(self):
        for item in self.db_tree.get_children():
            self.db_tree.delete(item)

    def show_context_menu(self, event):
        # Show context menu on right-click
        if not self.db_tree.selection():
            messagebox.showinfo("Info", "Please select a row first.")
        else:
            try:
                self.context_menu.tk_popup(event.x_root, event.y_root)
            finally:
                self.context_menu.grab_release()

    def edit_row(self):
        # Get the selected row
        selected_item = self.db_tree.selection()
        if not selected_item:
            messagebox.showinfo("Info", "Please select a row first.")
            return
        item = self.db_tree.item(selected_item)

        # Get current values
        current_values = item['values']
        new_values = self.edit_popup(current_values)
        if new_values:
            # Update the database
            self.cursor.execute('''
                UPDATE workbooks
                SET value = ?, quantity = ?, stacked = ?
                WHERE workbook_name = ? AND sheet_name = ? AND value = ?
            ''', (new_values[2], new_values[3], new_values[4], current_values[0], current_values[1], current_values[2]))
            self.conn.commit()
            self.update_table()

    def add_row(self):
        # Get the selected workbook and sheet name
        selected_workbook = self.workbook_selector.get()
        selected_sheet = self.sheet_selector.get()

        if selected_workbook == "--Select--" or selected_sheet == "--Select--":
            messagebox.showinfo("Info", "Please select a workbook and sheet first.")
            return

        # Get new values
        new_values = self.edit_popup([selected_workbook, selected_sheet, "", 0, 0], new_row=True)
        if new_values:
            # Insert into the database
            self.cursor.execute('''
                INSERT INTO workbooks (workbook_name, sheet_name, value, quantity, stacked)
                VALUES (?, ?, ?, ?, ?)
            ''', (new_values[0], new_values[1], new_values[2], new_values[3], new_values[4]))
            self.conn.commit()
            self.update_table()

    def edit_popup(self, values, new_row=False):
        popup = tk.Toplevel()
        popup.title("Edit Row" if not new_row else "Add Row")

        tk.Label(popup, text="Workbook:").grid(row=0, column=0, padx=10, pady=10)
        tk.Label(popup, text=values[0]).grid(row=0, column=1, padx=10, pady=10)

        tk.Label(popup, text="Sheet:").grid(row=1, column=0, padx=10, pady=10)
        tk.Label(popup, text=values[1]).grid(row=1, column=1, padx=10, pady=10)

        tk.Label(popup, text="Value:").grid(row=2, column=0, padx=10, pady=10)
        value_entry = tk.Entry(popup)
        value_entry.grid(row=2, column=1, padx=10, pady=10)
        value_entry.insert(0, values[2])

        tk.Label(popup, text="Quantity:").grid(row=3, column=0, padx=10, pady=10)
        quantity_entry = tk.Entry(popup)
        quantity_entry.grid(row=3, column=1, padx=10, pady=10)
        quantity_entry.insert(0, values[3])

        stacked_var = tk.IntVar(value=values[4])
        stacked_check = tk.Checkbutton(popup, text="Stacked", variable=stacked_var)
        stacked_check.grid(row=4, column=0, columnspan=2, pady=10)

        def on_submit():
            popup.new_values = [values[0], values[1], value_entry.get(), int(quantity_entry.get()), stacked_var.get()]
            popup.destroy()

        submit_button = tk.Button(popup, text="Submit", command=on_submit)
        submit_button.grid(row=5, column=0, columnspan=2, pady=10)

        popup.transient(self.root)
        popup.grab_set()
        self.root.wait_window(popup)

        return getattr(popup, 'new_values', None)

    def export_to_excel(self):
        # Open a file dialog to select the save location for the new Excel file
        save_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=(("Excel files", "*.xlsx"), ("All files", "*.*")),
            title="Save Excel File"
        )
        if save_path:
            # Create a new workbook
            new_wb = openpyxl.Workbook()
            new_wb.remove(new_wb.active)  # Remove the default sheet
            workbooks = set(row[0] for row in self.cursor.execute('SELECT DISTINCT workbook_name FROM workbooks').fetchall())
            for workbook_name in workbooks:
                sheets = set(row[0] for row in self.cursor.execute('SELECT DISTINCT sheet_name FROM workbooks WHERE workbook_name = ?', (workbook_name,)).fetchall())
                for sheet_name in sheets:
                    new_ws = new_wb.create_sheet(title=sheet_name)
                    rows = self.cursor.execute('SELECT value, quantity, stacked FROM workbooks WHERE workbook_name = ? AND sheet_name = ?', (workbook_name, sheet_name)).fetchall()
                    for value, quantity, stacked in rows:
                        for _ in range(quantity):
                            if stacked:
                                value = value.replace(" ", " " * 10)  # Ensure 10 spaces between stacked items
                            new_ws.append([value])
            # Save the new workbook
            new_wb.save(save_path)
            messagebox.showinfo("Info", "Data successfully exported to Excel")

if __name__ == "__main__":
    root = tk.Tk()
    app = PSCXL(root)
    root.mainloop()
